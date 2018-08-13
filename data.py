import pandas as pd
#import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook
import numpy as np
from math import sqrt, pow, fabs
from scipy.stats import spearmanr


class Abundances():
    def __init__(self):
        self.abundance_df = pd.DataFrame(index=[], columns=[])
        self.corr_matrix = None
        self.corr_signature = None
        self.sample_names = []
        self.header_present = False
        self.abundance_raw_df = None
    
    def addMasking(self):
        """ merges abundance dataframe and taxonomy dataframe """
        self.abundance_df['masked'] = [False]*len(self.abundance_df.index)
        self.abundance_df['colour'] = ['undefined']*len(self.abundance_df.index)
        
        if(0):
            test_df = self.abundance_df[self.sample_names+self.tax_levels]
            test_df['id'] = self.abundance_df.index
            test_df = test_df.loc[:,['id'] + self.sample_names]
            test_df.index = self.abundance_df[self.tax_levels[0]]
            test_df.to_csv('/Users/klincke/Documents/MicroWine/GraphicsProject/Mac/abundance_table_conet_2016_11.tab', sep='\t')
    
    #old
    def addSample_old(self, sample_name, filename):
        """ adds a sample (as one column) to the dataframe """
        tax_levels = None
        if len(self.abundance_df.columns) == 0:
            #abundance_df = pd.read_table(filename, header=None, index_col=0) #krona (no header, no index)
            self.abundance_df = pd.read_table(filename, header=0, index_col=0)
            #try:
            #    int(abundance_df.values[0,0])
            #except:
            #    abundance_df.columns = abundance_df.iloc[0]
            #    abundance_df = abundance_df[1:]
            self.tax_levels = self.abundance_df.columns.tolist()
            #self.tax_levels = abundance_df.columns.tolist()
            self.tax_levels = self.tax_levels[1:]
            self.abundance_df = self.abundance_df.rename(columns = {list(self.abundance_df.columns)[0]: sample_name})
                #abundance_df = abundance_df.rename(columns = {list(abundance_df.columns)[0]: sample_name})
                
            self.abundance_df = self.abundance_df[self.tax_levels + [sample_name]]
            #self.abundance_df = abundance_df[self.tax_levels + [sample_name]]
            self.abundance_df.index = self.abundance_df['species']
            #self.abundance_df.index = self.abundance_df[self.tax_levels[0]]
            tax_levels = list(self.abundance_df.columns)
            tax_levels.remove(sample_name)
        else:
            sample_df = pd.read_table(filename, header=0, index_col=0)
            sample_df = sample_df.rename(columns = {'abundance': sample_name})
            #sample_df.index = sample_df['species']
            sample_df.index = sample_df[self.tax_levels[0]]
            self.abundance_df = pd.merge(self.abundance_df, sample_df, how='outer', on=tax_levels)
            self.abundance_df.fillna(value=0, inplace=True)
        self.sample_names.append(sample_name.strip())
        #print(self.abundance_df.columns)
        #print(self.abundance_df.head(2))
        return self.tax_levels

    #add MGmapper sample (with raw counts)
    def addMGmapperSample(self, sample_name, filename):
        """ adds a sample (as one column) to the dataframes for relative and raw counts"""
        tax_levels = None
        if len(self.abundance_df.columns) == 0:
            self.abundance_df = pd.read_table(filename, header=0) #krona (no header, no index)
            cols = list(self.abundance_df.columns)
            self.abundance_df = self.abundance_df[cols[0:2] + cols[:2:-1]]
            self.tax_levels = self.abundance_df.columns.tolist()[2:]
            self.abundance_df = self.abundance_df[self.abundance_df.columns.tolist()[0:2] + self.tax_levels]
            self.abundance_df.rename(columns={self.abundance_df.columns[0]:sample_name}, inplace=True)
            self.abundance_df.set_index(self.tax_levels[0], drop=False, inplace=True)
            
            self.abundance_raw_df = self.abundance_df.loc[:,[self.abundance_df.columns[1]] + self.tax_levels]
            self.abundance_raw_df.rename(columns={self.abundance_raw_df.columns[0]:sample_name}, inplace=True)
            self.abundance_df = self.abundance_df.loc[:,[self.abundance_df.columns[0]] + self.tax_levels]
        else:
            sample_df = pd.read_table(filename, header=0)
            sample_raw_df = sample_df.loc[:,[sample_df.columns[1]]+self.tax_levels]
            sample_raw_df.rename(columns={sample_raw_df.columns[0]:sample_name}, inplace=True)  
            sample_raw_df.set_index(self.tax_levels[0], drop=False, inplace=True)
            sample_df.rename(columns={sample_df.columns[0]:sample_name}, inplace=True)  
            sample_df.set_index(self.tax_levels[0], drop=False, inplace=True)
            self.abundance_df = pd.merge(self.abundance_df, sample_df, how='outer', on=self.tax_levels)
            self.abundance_df.set_index(self.tax_levels[0], drop=False, inplace=True)   
            self.abundance_df.fillna(value=0, inplace=True) 
            self.abundance_raw_df = pd.merge(self.abundance_raw_df, sample_raw_df, how='outer', on=self.tax_levels)
            self.abundance_raw_df.set_index(self.tax_levels[0], drop=False, inplace=True)   
            self.abundance_raw_df.fillna(value=0, inplace=True)
        self.abundance_df[sample_name] = self.abundance_df[sample_name].astype(float)
        self.abundance_raw_df[sample_name] = self.abundance_raw_df[sample_name].astype(float)
        
        self.sample_names.append(sample_name.strip())
        self.abundance_df = self.abundance_df[self.sample_names + self.tax_levels]
        self.abundance_raw_df = self.abundance_raw_df[self.sample_names + self.tax_levels]
        myindex = list(self.abundance_df.index)
        newlist =  sorted(set([i for i in myindex if myindex.count(i)>1]))
        #problems with the ncbi taxonomy (typos?)
        for i in newlist:
            self.abundance_df.loc[i,self.sample_names] = self.abundance_df.loc[i].sum(numeric_only=True)
            self.abundance_df.drop(i, inplace=True)
            self.abundance_raw_df.loc[i,self.sample_names] = self.abundance_raw_df.loc[i].sum(numeric_only=True)
            self.abundance_raw_df.drop(i, inplace=True)
        return self.tax_levels

    #new version, to take also krona input    
    def addSample(self, sample_name, filename):
        """ adds a sample (as one column) to the dataframe """
        tax_levels = None
        #self.header_present = False
        if len(self.abundance_df.columns) == 0:
            self.abundance_df = pd.read_table(filename, header=None)#, index_col=0) #krona (no header, no index)
            #print(self.abundance_df.columns)
            #cols = self.abundance_df.columns
            #self.abundance_df = self.abundance_df[cols[0]+cols[1::-1]]
            #print(cols)
            #print(self.abundance_df.columns)
            try:    #check if header is in file
                int(self.abundance_df.values[0,0])
            except ValueError:
                self.header_present = True
                #print self.abundance_df.columns
                self.abundance_df.columns = self.abundance_df.iloc[0]
                self.abundance_df = self.abundance_df[1:]
            #print(self.abundance_df.columns)
            cols = list(self.abundance_df.columns)
            self.abundance_df = self.abundance_df[[cols[0]] + cols[:0:-1]]
            self.tax_levels = self.abundance_df.columns.tolist()[1:]
            #print('cols:')
            #print(cols)
            #print(self.abundance_df.columns)
            self.abundance_df = self.abundance_df[[self.abundance_df.columns.tolist()[0]] + self.tax_levels]
            if not self.header_present: 
                self.abundance_df = self.abundance_df.loc[:,[0] + self.tax_levels[::-1]]
                self.abundance_df.rename(columns=dict(zip(self.abundance_df.columns[1:], self.tax_levels)),inplace=True)
            self.abundance_df.rename(columns={self.abundance_df.columns[0]:sample_name}, inplace=True)
            #print(self.tax_levels[0])
            self.abundance_df.set_index(self.tax_levels[0], drop=False, inplace=True)
            #self.abundance_df.reindex(self.tax_levels[0])
            #print(self.abundance_df.index)
        else:
            #print(self.header_present)
            sample_df = pd.read_table(filename, header=None)#, index_col=0)
            if self.header_present:
                sample_df.columns = sample_df.iloc[0]
                sample_df = sample_df[1:]
            else:
                #sample_df = sample_df.loc[:,[0] + sample_df.columns.tolist()[1:][::-1]]
                sample_df = sample_df.loc[:,[0] + self.tax_levels[::-1]]
                #sample_df.rename(columns=dict(zip(sample_df.columns[1:], sample_df.columns.tolist()[1:])),inplace=True)
            #sample_df.rename(columns={sample_df.columns[0]:sample_name}, inplace=True)
                sample_df.rename(columns=dict(zip(sample_df.columns[1:], self.tax_levels)),inplace=True)
            #print(sample_df.head(2))
            sample_df.rename(columns={sample_df.columns[0]:sample_name}, inplace=True)  
            sample_df.set_index(self.tax_levels[0], drop=False, inplace=True)
            #sample_df.reindex(self.tax_levels[0])
            self.abundance_df = pd.merge(self.abundance_df, sample_df, how='outer', on=self.tax_levels)
            self.abundance_df.set_index(self.tax_levels[0], drop=False, inplace=True)   
            self.abundance_df.fillna(value=0, inplace=True) 
        self.abundance_df[sample_name] = self.abundance_df[sample_name].astype(float)
        
        
        self.sample_names.append(sample_name.strip())
        self.abundance_df = self.abundance_df[self.sample_names + self.tax_levels]
        #for c in self.abundance_df.columns:
        #    print(c)
        #    print(self.abundance_df[c].dtype)
        #print(len(self.abundance_df.index))
        #print(len(set(list(self.abundance_df.index))))
        myindex = list(self.abundance_df.index)
        newlist =  sorted(set([i for i in myindex if myindex.count(i)>1]))
        #problems with the ncbi taxonomy (typos?)
        for i in newlist:
            self.abundance_df.loc[i,self.sample_names] = self.abundance_df.loc[i].sum(numeric_only=True)
            self.abundance_df.drop(i, inplace=True)
        return self.tax_levels
    
    def addMetaPhlanSample(self, sample_name, filename):
        #tax_levels = None
        #self.header_present = False
        taxonomy_keys = ['s', 'g', 'f', 'o', 'c', 'p', 'k']
        if len(self.abundance_df.columns) == 0:
            self.abundance_df = pd.DataFrame()
            header = False
            header_list = None
            with open(filename) as infile:
                for line in infile:
                    if line[0] != '#':
                        #for item in ['k__', 'p__', 'c__', 'o__', 'f__', 'g__', 's__']:
                        if '|s__' in line:
                            if 'd__' in line:
                                line = line.replace('d__', 'k__')
                            #print(line)
                        #if '|s__' in line and 'g__' in line and 'f__' in line and 'o__' in line and 'c__' in line and 'p__' in line and 'k__' in line:
                            taxonomy, abundance = line.strip().split('\t')
                            taxonomy_series = pd.Series(np.array(['-']*7), index=taxonomy_keys)
                            #print(taxonomy)
                            for pair in taxonomy.split('|'):
                                key, value = pair.split('__')
                                
                                try:
                                    taxonomy_series[key] = value
                                except KeyError:
                                    pass
                            taxonomy_list = list(taxonomy_series)
                            if not header:
                                header_dict = {'s':'species', 'g':'genus', 'f':'family', 'o':'order', 'c':'class', 'p':'phylum', 'k':'kingdom'}
                                header_list = [header_dict[item.split('__')[0]] for item in taxonomy.split('|')][::-1]
                                header_list = ['species', 'genus', 'family', 'order', 'class', 'phylum', 'kingdom']
                                header = True
                                self.abundance_df = pd.DataFrame(columns = [sample_name] + header_list)
                            #print(taxonomy_list)
                            self.abundance_df.loc[taxonomy_list[0]] = [float(abundance)] + taxonomy_list
            self.tax_levels = header_list
        else:
            sample_df = pd.DataFrame(columns=[sample_name] + self.tax_levels)
            with open(filename) as infile:
                for line in infile:
                    if line[0] != '#':
                        if '|s__' in line:
                            if 'd__' in line:
                                line = line.replace('d__', 'k__')
                        #if '|s__' in line and 'g__' in line and 'f__' in line and 'o__' in line and 'c__' in line and 'p__' in line and 'k__' in line:
                            taxonomy, abundance = line.strip().split('\t')
                            taxonomy_series = pd.Series(np.array(['-']*7), index=taxonomy_keys)
                            for pair in taxonomy.split('|'):
                                key, value = pair.split('__')
                                try:
                                    taxonomy_series[key] = value
                                except KeyError:
                                    pass
                            taxonomy_list = list(taxonomy_series)
                            sample_df.loc[taxonomy_list[0]] = [float(abundance)] + taxonomy_list
            self.abundance_df = pd.merge(self.abundance_df, sample_df, how='outer', on=self.tax_levels)
            self.abundance_df.set_index(self.tax_levels[0], drop=False, inplace=True)   
            self.abundance_df.fillna(value=0, inplace=True) 
        self.abundance_df[sample_name] = self.abundance_df[sample_name].astype(float)
    
        self.sample_names.append(sample_name.strip())
        self.abundance_df = self.abundance_df[self.sample_names + self.tax_levels]
    
        #print(self.abundance_df.head(2))
        #print(self.abundance_df)
        return self.tax_levels
    
    
    # #takes krona input (without header)
    # def addSample(self, sample_name, filename):
    #     """ adds a sample (as one column) to the dataframe """
    #     tax_levels = None
    #     #self.header_present = False
    #     if len(self.abundance_df.columns) == 0:
    #         self.abundance_df = pd.read_table(filename, header=None)#, index_col=0) #krona (no header, no index)
    #         try:    #check if header is in file
    #             int(self.abundance_df.values[0,0])
    #         except ValueError:
    #             self.header_present = True
    #             self.abundance_df.columns = self.abundance_df.iloc[0]
    #             self.abundance_df = self.abundance_df[1:]
    #         self.tax_levels = self.abundance_df.columns.tolist()[1:]
    #         self.abundance_df = self.abundance_df[[self.abundance_df.columns.tolist()[0]] + self.tax_levels]
    #         if not self.header_present:
    #             self.abundance_df = self.abundance_df.loc[:,[0] + self.tax_levels[::-1]]
    #             self.abundance_df.rename(columns=dict(zip(self.abundance_df.columns[1:], self.tax_levels)),inplace=True)
    #         self.abundance_df.rename(columns={self.abundance_df.columns[0]:sample_name}, inplace=True)
    #         #print(self.tax_levels[0])
    #         self.abundance_df.set_index(self.tax_levels[0], drop=False, inplace=True)
    #         #self.abundance_df.reindex(self.tax_levels[0])
    #         #print(self.abundance_df.index)
    #     else:
    #         #print(self.header_present)
    #         sample_df = pd.read_table(filename, header=None)#, index_col=0)
    #         if self.header_present:
    #             sample_df.columns = sample_df.iloc[0]
    #             sample_df = sample_df[1:]
    #         else:
    #             #sample_df = sample_df.loc[:,[0] + sample_df.columns.tolist()[1:][::-1]]
    #             sample_df = sample_df.loc[:,[0] + self.tax_levels[::-1]]
    #             #sample_df.rename(columns=dict(zip(sample_df.columns[1:], sample_df.columns.tolist()[1:])),inplace=True)
    #         #sample_df.rename(columns={sample_df.columns[0]:sample_name}, inplace=True)
    #             sample_df.rename(columns=dict(zip(sample_df.columns[1:], self.tax_levels)),inplace=True)
    #         #print sample_df.head(2)
    #         sample_df.rename(columns={sample_df.columns[0]:sample_name}, inplace=True)
    #         sample_df.set_index(self.tax_levels[0], drop=False, inplace=True)
    #         #sample_df.reindex(self.tax_levels[0])
    #         #print('addSample')
    #         #print(sample_df.head(2))
    #         #print(sample_df.index)
    #         self.abundance_df = pd.merge(self.abundance_df, sample_df, how='outer', on=self.tax_levels)
    #     self.abundance_df[sample_name] = self.abundance_df[sample_name].astype(float)
    #     #for c in self.abundance_df.columns:
    #     #    print(self.abundance_df[c].dtype)
    #     self.abundance_df.fillna(value=0, inplace=True)
    #     self.sample_names.append(sample_name.strip())
    #     #print(self.abundance_df.head(2))
    #     #print(self.abundance_df)
    #     return self.tax_levels
            

    def deselectOfSample(self, names, current_tax_level):
        """ deselects/removes species given the name of the species """
        index_list = list(self.sample[self.sample['masked'] == False].index)
        for name in names:
            idx = self.sample[self.sample[current_tax_level]==name].index
            self.sample.at[idx, 'masked'] = True

    def getAbundanceForOneSample(self, species, sample):
        """ gets abundance for one species of one sample """
        return self.abundance_df.loc[species,sample]

    def getAbundances(self, species):
        """ gets the abundances in all samples for a given species (if it is unmarked) """
        #return self.abundance_df.loc[[species],:]    
        return self.abundance_df[self.abundance_df['species']==species]
    
    def getDataframe(self):
        return self.abundance_df
    
    
    def get_corr(self):
        """ gets the correlation matrix and the signature(sample names) """
        return self.corr_matrix, self.corr_signature

    
    def getCorrelationForSpecies(self, current_species, threshold):
        """ calculates correlation for abundances of species (between species) """
        grouped = self.groupAllSamples()
        if self.corr_matrix is None or self.corr_signature is None or self.corr_signature[0] != grouped.iloc[:,len(self.tax_levels):-1].columns.tolist():
            corr_matrix = grouped.iloc[:,len(self.tax_levels):-2]
            corr_matrix.index = grouped[self.tax_level]
            self.corr_matrix = corr_matrix.transpose().corr(method='spearman')
            self.corr_signature = (list(corr_matrix.columns), self.tax_levels[0]) 

        corr_matrix = self.corr_matrix.loc[:,current_species]
        text = 'spearman (rank) correlation >= ' + str(threshold) + ':\n'

        corr_series =  corr_matrix[abs(corr_matrix) >= threshold].sort_values(ascending=False)
        
        corr_matrix = grouped.iloc[:,len(self.tax_levels):-1]
        corr_matrix.index = grouped[self.tax_level]
        corr_list = []

        current_abundance = corr_matrix.loc[current_species,:corr_matrix.columns[-2]]
        list_index = []
        for name in corr_matrix.index:
            new_abundance = corr_matrix.loc[name,:][:-1]
            corr = '{0:.3}'.format(current_abundance.corr(new_abundance, method='spearman'))
            if corr != 'nan' and abs(float(corr)) >= threshold and current_species != name:
                corr_list.append('{0:.3}'.format(current_abundance.corr(new_abundance, method='spearman')))
                list_index.append(name)
            #rho, pval = spearmanr(current_abundance, new_abundance)
            #if rho != 'nan' and abs(float(corr)) >= threshold and current_species != name and pval <= 0.05:
            #    #corr_list.append('{0:.3}'.format(rho))
            #    #list_index.append(name)
            #    print(name + '\t' + str(rho) + '\t' + str(pval))
            
        #for i in xrange(len(corr_list)):
        #    if corr_list[i] != 'nan' and abs(float(corr_list[i])) >= threshold and current_species != corr_matrix.index[i]:
        #        print(corr_matrix.index[i] + '\t' + corr_list[i])
        corr_series = pd.Series(corr_list, index=list_index)
        return text, corr_series
        #return text, corr_series.map('{0:.3}'.format)
    
    
    def corr(self, tax_levels , samples, tax_level, superfamily_groups, min_samples):
        """ correlation """
        corr_dict = {}
        corr_df = pd.DataFrame(columns=('name1', 'name2', 'r', 'p', 'max_name1', 'max_name2'))
        tax_level_idx = tax_levels.index(tax_level)
        taxlevelnum = len(tax_levels) - tax_level_idx + 1
        samplenum = len(samples)
        #grouped1 = self.abundance_df.groupby(tax_levels[tax_level_idx:], sort=False, as_index=False).sum()
        #print(self.abundance_df[self.abundance_df['masked']==False].head)
        grouped1 = self.abundance_df[self.abundance_df['masked']==False].groupby(tax_levels[tax_level_idx:], sort=False, as_index=False).sum()
        #for i, idx in enumerate(grouped1.index):
        #    if grouped1.iloc[i,taxlevelnum:-1].astype(bool).sum() < min(samplenum * 0.2, 3) :
        #        grouped1.at[i, 'masked'] = True
        #grouped1_filtered = grouped1[grouped1['masked'] == False]
        #print(len(grouped1_filtered.index))
        #group1 = grouped1_filtered
        group1 = grouped1[grouped1[tax_levels[-1]] == superfamily_groups[0]]
        group2 = grouped1[grouped1[tax_levels[-1]] == superfamily_groups[1]]
        group1 = group1.loc[:,tax_levels + samples]
        group2 = group2.loc[:,tax_levels + samples]
        
        k = 0
        if superfamily_groups[0] == superfamily_groups[1]:
            for i, idx in enumerate(group1.index):
                abundance1 = group1.iloc[i, taxlevelnum:-1]#-2
                if len(abundance1[abundance1 > 0.00]) < min_samples:
                    #
                    group2.drop(idx, inplace=True)
                    continue
                name1 = group1.loc[idx, tax_level]
                if name1 != '-':
                    max_name1 = max(abundance1)
                    for j in xrange(i+1, len(group2.index)):
                        jdx = group2.index[j]
                        abundance2 = group2.iloc[j, taxlevelnum:-1]
                        if len(abundance2[abundance2 > 0.00]) < min_samples:
                            continue
                        name2 = group2.loc[jdx, tax_level]
                        if name2 != '-':
                            #rho, pval = spearmanr(abundance1, abundance2)
                            max_name2 = max(abundance2)
                            rho, pval = spearmanr(abundance1, abundance2)
                            corr_df.loc[k] = [name1, name2, rho, pval, max_name1, max_name2]
                            corr_dict[(name1 ,name2)] = {}
                            corr_dict[(name1 ,name2)]['r'] = rho
                            corr_dict[(name1 ,name2)]['p'] = pval
                            k += 1
        else:
            for i, idx in enumerate(group1.index):
                abundance1 = group1.iloc[i, taxlevelnum:]
                if len(abundance1[abundance1 > 0.00]) < min_samples:
                    #if len(abundance1[abundance1 > 0.00]) == 0:
                    #    group2.drop([idx], inplace=True)
                    #    continue
                    #group2.drop(idx, inplace=True)
                    continue
                #abundance1[self.sample_names] > 0) * 1
                name1 = group1.loc[idx, tax_level]
                if name1 != '-':
                    max_name1 = max(abundance1)
                    for j, jdx in enumerate(group2.index):
                        abundance2 = group2.iloc[j, taxlevelnum:]
                        if len(abundance2[abundance2 > 0.00]) < min_samples:
                            continue
                        name2 = group2.loc[jdx, tax_level]
                        if name2 != '-':
                            #rho, pval = spearmanr(abundance1, abundance2, max_name1, max_name2)
                            max_name2 = max(abundance2)
                            rho, pval = spearmanr(abundance1.values, abundance2.values)
                            corr_df.loc[k] = [name1, name2, rho, pval, max_name1, max_name2]
                            corr_dict[(name1 ,name2)] = {}
                            corr_dict[(name1 ,name2)]['r'] = rho
                            corr_dict[(name1 ,name2)]['p'] = pval
                            k += 1
                            #if fabs(rho) > float(r) and pval < float(p):
                            #    print(name1 + '\t' + name2 + '\t' + str(rho) + '\t' + str(pval))
        return corr_df

    def getMaxAbundanceOfSample(self):
        """ gets the maximum abundance of all unmasked species """
        try: maximum = max(self.sample[self.sample['masked'] == False]['abundance'])+0.01
        except: maximum = 0
        return maximum
    
    def getPresenceAbsenceDF(self, threshold):
        """ gets a dataframe giving the presence/absence of organisms """
        binaryAbundance = (self.abundance_df[self.sample_names] > threshold) * 1
        binaryAbundance.index = list(self.abundance_df['species'])
        return binaryAbundance
    
    def getSample(self, sample_name):
        """ gets sample to work with """
        columns = self.tax_levels + [sample_name, 'masked']
        self.sample = self.abundance_df[columns]
        self.sample = self.sample[self.sample[sample_name] > 0]
        self.sample = self.sample.rename(columns = {sample_name: 'abundance'})
        self.sample[self.sample['masked']==False]
        
    def getSamplesList(self):
        """ gets the list of samples (currently loaded) """
        return self.sample_names    
    
    def getWorkingSample(self, tax_level, as_index=False):
        """ gets sample grouped on current tax_level """
        levels = self.tax_levels[self.tax_levels.index(tax_level):]
        grouped = self.sample.groupby(levels, sort=False, as_index=as_index).sum()#['abundance']
        grouped.index = range(len(grouped.index))
        self.tax_level = tax_level
        return grouped[grouped['masked'] == False]
       
    def getValuesForColumn(self, columnname):
        """ gets as list of the unique entries of a column"""
        return list(self.abundance_df[columnname].unique())
       
    def getNotHidden(self):
        """ gets a dataframe of containing only the rows which are not masked/hidden """
        return self.abudance_df[self.abundance_df['maske']==False]
       
    def groupAllSamples(self, all_levels=None):
        """ groups all samples on tax_level """
        if all_levels is None:
            levels = self.tax_levels[self.tax_levels.index(self.tax_level):]
        else:
            levels = all_levels
        grouped = self.abundance_df.groupby(levels, sort=False, as_index=False).sum()
        if self.tax_level == self.tax_levels[0]:
            grouped['colour'] = list(self.abundance_df['colour'])
        else:
            grouped['colour'] = ['undefined']*len(grouped.index)
        grouped.index = range(len(grouped.index))
        return grouped[grouped['masked'] == False]
       
    def groupAbsoluteSamples(self):
        """ groups absolute abundances """
        unmasked_tax = set(list(self.groupAllSamples(all_levels=self.tax_levels)[self.tax_level]))
        levels = self.tax_levels[self.tax_levels.index(self.tax_level):]
        if self.abundance_raw_df is not None:
            grouped = self.abundance_raw_df.groupby(levels, sort=False, as_index=False).sum()
            grouped.index = grouped[self.tax_level]
            #grouped.index = range(len(grouped.index))
            #return grouped[grouped['masked'] == False]
            return grouped[grouped[self.tax_level].isin(unmasked_tax)]
            #return grouped
        else:
            return None
       
    def group_binary_patterns(self):
        """ creates a binary pattern """
        binary_abundance_df = self.getPresenceAbsenceDF(0)
        binary_abundance_df.sort_values(by=list(binary_abundance_df.columns))
        pattern_dict= {}
        for idx1 in binary_abundance_df.index:
            pattern = ' '.join(map(str, list(binary_abundance_df.loc[idx1,:])))
            if pattern not in pattern_dict:
                pattern_dict[pattern] = []
            pattern_dict[pattern].append(idx1)
        return pattern_dict
    
    def reset(self):
        """ resets options for displaying the graph (includes all species 
        that are in the sample) """
        self.sample['masked'] = [False]*len(self.sample.index)
        self.sample['colour'] = ['undefined']*len(self.sample.index)
        
    def selectOfSample(self, indexes):
        """ selects species given the indexes"""
        index_set = set()
        for idx in indexes:
            i = list(self.sample[self.sample['masked'] == False].index)[idx]
            index_set.add(i)
        for ind in list(self.sample[self.sample['masked'] == False].index):
            if ind not in index_set:
                self.sample.at[ind, 'masked'] = True
        return index_set
        
    def set_corr(corr_matrix, corr_signature):
        """ sets correlation variables """
        self.corr_matrix = corr.matrix
        self.corr_signature = corr.signature    
    
    def sortSample(self, key, ascending):
        """ sorts sample by key """
        try:
            self.sample[self.sample['masked'] == False].sort_values(by=key, ascending=ascending)
        except:
            pass
        
    def renewMasking(self, indices, colours_dict):
        """ renews the masking and colours """
        for idx in self.abundance_df.index:
            if idx in indices:
                self.abundance_df.loc[idx, 'masked'] = False
            else:
                self.abundance_df.loc[idx, 'masked'] = True
            if idx in colours_dict:
                self.abundance_df.loc[idx, 'colour'] = colours_dict[idx]
            else: 
                self.abundance_df.loc[idx, 'colour'] = 'undefined'
        
    def randomForestClassifier(self, train_cols, test_cols, targets, feature_selction_var, min_abundance_threshold, shuffle=False):
        """ run random forest classification """
        from sklearn.ensemble import RandomForestClassifier
        #from sklearn.ensemble import RandomForestRegressor
        
        #train = self.abundance_df.loc[:,train_cols] #train.as_matrix(cols)
        train = self.abundance_df[self.abundance_df['masked']==False].loc[:,train_cols] #train.as_matrix(cols)
        #test = self.abundance_df.loc[:,test_cols] #.as_matrix(test_cols)
        test = self.abundance_df[self.abundance_df['masked']==False].loc[:,test_cols] #.as_matrix(test_cols)
        #names = list(self.abundance_df.loc[:, 'species'])
        names = list(self.abundance_df[self.abundance_df['masked']==False].loc[:, 'species'])
        
        #most_common_species_set = set()
        #for col in train_cols:
        #    sorted_series = self.abundance_df.loc[:, col].sort_values(ascending=False)[:100]
        #    most_common_species_set |= set(list(sorted_series.index))
        #most_common_species_list = []
        #for id0 in most_common_species_set:
        #    #print(max(self.abundance_df.loc[id0,train_cols]))
        #    if max(self.abundance_df.loc[id0,train_cols]) >= min_abundance_threshold:
        #        most_common_species_list.append(id0)
        ##print(len(most_common_species_list))
        #most_common_species_set = set(most_common_species_list)
        #train = train.loc[list(most_common_species_set),:]
        #test = test.loc[list(most_common_species_set),:]
        #names = list(self.abundance_df.loc[list(most_common_species_set),'species'])
        
        #feature selection by variance
        from sklearn.feature_selection import VarianceThreshold
        sel = VarianceThreshold(threshold=(0.999 * (1 - 0.999)))  
        if feature_selction_var:
            #ds1 = np.transpose(ds10.as_matrix())
            #ds1 = sel.fit_transform(np.transpose(ds10.as_matrix()))
            #ds2 = np.transpose(ds20.as_matrix())
            #train = sel.fit_transform(np.transpose(train.as_matrix()))
            train = sel.fit_transform(np.transpose(train.values))
            
            #names = list(self.abundance_df.loc[:, 'species'].as_matrix()[sel.get_support()])
            #names = list(self.abundance_df[self.abundance_df['masked']==False].loc[:, 'species'].as_matrix()[sel.get_support()])
            names = list(self.abundance_df[self.abundance_df['masked']==False].loc[:, 'species'].values[sel.get_support()])
            #test = sel.fit_transform(np.transpose(test.as_matrix()))
            test = sel.fit_transform(np.transpose(test.values))
            ds10 = np.asmatrix(train)[[i for i, j in enumerate(targets) if j == 0],:]
            ds1 = np.transpose(sel.fit_transform(np.transpose(ds10)))
        else:

            #train = np.transpose(train.as_matrix())
            train = np.transpose(train.values)
            #test = np.transpose(test.as_matrix())
            test = np.transpose(test.values)
            ds10 = train.iloc[:,[i for i, j in enumerate(targets) if j == 0]]
            #ds1 = np.transpose(ds10.as_matrix())
            ds1 = np.transpose(ds10.values)

        if shuffle == 'index':
            from random import shuffle
            shuffle(names)

        #rf = RandomForestClassifier(n_estimators=10)
        target = targets 
        #group1 = list(self.abundance_df.loc[:,train_cols].columns[:target.count(0)])
        group1 = list(self.abundance_df[self.abundance_df['masked']==False].loc[:,train_cols].columns[:target.count(0)])
        #group2 = list(self.abundance_df.loc[:,train_cols].columns[target.count(0):])
        group2 = list(self.abundance_df[self.abundance_df['masked']==False].loc[:,train_cols].columns[target.count(0):])

        #rf = RandomForestRegressor(n_estimators=1000)#, class_weight="balanced")
        rf = RandomForestClassifier(n_estimators=1000)  # bootstrap=False
        #, max_features=100)#, min_sample_leaf=50)
        #rf = RandomForestRegressor(n_estimators=20, max_features=2)
        #class_weight="balanced" #{class_label: weight}
        #n_estimators=1000,
        rf.fit(train, target)
        
        #from sklearn.metrics import roc_auc_score
        #for l in leaf:
        #model = RandomForestRegressor(min_samples_split=2, max_depth=None, bootstrap=False, min_samples_leaf=2)
        #    #n_estimator=200, oob_score=True, min_samples_leaf=10,max_features=f, 
        #model.fit(train,target)
        #    #print("AUC - ROC : ")
        #    #print(roc_auc_score(target,model.oob_prediction_))
        #    #print(model.feature_importances_)
            
        #from sklearn.ensemble import ExtraTreesClassifier
        #model = ExtraTreesClassifier()
        #model.fit(train, target)
        
        from treeinterpreter import treeinterpreter as ti
        prediction, bias, contributions = ti.predict(rf, np.array(train))
        
        #for i in range(len(train)):
        #    j = 0
        # #   print(i)
        #    #print("\tBias (trainset mean)")
        #    #print(bias[i])
        # #   print(contributions[0][0])
        #    #for c, feature in sorted(zip(contributions[i], 
        #    #                             names), 
        #    #                            #self.abundance_df.index), 
        #    #                         key=lambda x: -abs(x[0])):
        #    for c, feature in zip(contributions[i], list(self.abundance_df.index)):
        #        if c[0] != 0:
        #        #print feature, ':\t', "{:.2e}".format(c), '\t', self.abundance_df.loc[feature, 'species']
        #            if j <10:
        #  #              print()'\t' + self.abundance_df.loc[feature, 'species'], '\t', "{:.2e}".format(c[0]))
        #                j += 1
        totalc = np.mean(contributions, axis=0)    
        
        #from sklearn import model_selection
        #from sklearn.model_selection import cross_val_score
        #clf = RandomForestClassifier(n_estimators=10, max_depth=None, min_samples_split=2, random_state=0)
        #scores = cross_val_score(clf, X, y)
        
        ##compare 2 groups of samples
        prediction1, bias1, contributions1 = ti.predict(rf, np.array(ds1))

        mean_contri = [0 for i in xrange(len(names))]
        for s in xrange(len(ds1)):
            for i in xrange(len(names)):
                mean_contri[i] += contributions1[s][i][0]
        mean_contri = [x/len(ds1)for x in mean_contri]
            
        names_list = []
        #for c, org in sorted(zip(mean_contri, list(self.abundance_df.loc[:,'species'])), reverse=True):
        for c, org in sorted(zip(mean_contri, names), reverse=True):
            if c != 0:
                #print(self.abundance_df.loc[i,group1])
                #idx = self.abundance_df[self.abundance_df['species'] == org].index.tolist()[0]
                idx = self.abundance_df[self.abundance_df['masked']==False][self.abundance_df['species'] == org].index.tolist()[0]
                if shuffle:
                    #print(names.index(org))
                    #idx = list(self.abundance_df.index)[names.index(org)]
                    idx = list(self.abundance_df[self.abundance_df['masked']==False].index)[names.index(org)]
                #maximum = max(self.abundance_df.loc[idx,group1 + group2])
                maximum = max(self.abundance_df[self.abundance_df['masked']==False].loc[idx,group1 + group2])
                #print(str(round(c, 3)) + '\t' + org + '\t' + str(round(maximum,3)))
                names_list.append([round(c, 3), org, round(maximum,3)])
        
        return names_list      
        
    def shape(self):
        """ gets the shape(dimensions) of the dataframe """
        return self.abundance_df.shape
        
    def save_count_tables(self):
        """  """
        from pathlib import Path
        path = Path(__file__).parent
        #print(str(path) + '/relative_counts.csv')
        if self.abundance_df is not None:
            self.abundance_df[self.sample_names].to_csv(str(path) + '/relative_counts.csv')
        if self.abundance_raw_df is not None:
            self.abundance_raw_df[self.sample_names].to_csv(str(path) + '/absolute_counts.csv')
        
    
        
class MetaData():
    #"class to "
    def __init__(self, filename, sample_name, abundance_df):
        self.abundance_df = abundance_df
        #wb = openpyxl.load_workbook(filename)
        wb = load_workbook(filename)
        for name in wb.get_sheet_names():
            sheet = wb.get_sheet_by_name(name)
            #if sheet['B1'].value.upper() in sample_name.upper():
            sample_names_metadata = []
            #for cell in sheet.columns[1]:
            for cell in list(wb.active.columns)[1]:
                sample_names_metadata.append(cell.value)
            meta_dict = {}
            for colnum in xrange(3, sheet.max_column+1):
                key = sheet.cell(row=1, column=colnum).value.upper()
                meta_dict[key] = []
                for rownum in xrange(2, sheet.max_row+1):
                    meta_dict[key].append(sheet.cell(row=rownum, column=colnum).value)
            self.meta_df = pd.DataFrame(meta_dict, index=sample_names_metadata[1:])
            
    def getCorrelationForMetaData(self, name, tax_level):
        """ correlates species abundance with meta data """
        grouped = self.abundance_df.groupAllSamples()
        grouped.set_index(tax_level, drop=False, inplace=True) 
        samplenames = self.abundance_df.getSamplesList()
        grouped_series = grouped.loc[name,samplenames]
        meta_samples_df = pd.DataFrame(columns=[], index=[])
        for sample in grouped_series.index:
            for sample_meta in self.meta_df.index:
                if sample_meta.upper() in sample.upper():
                    meta_samples_df[sample] = list(self.meta_df.loc[sample_meta,:])
        meta_samples_df.index = self.meta_df.columns

        corr_list = []
        for item in meta_samples_df.index:
            corr_coef = meta_samples_df.loc[item,:].corr(grouped_series, method='spearman')   
            corr_list.append('{0:.3}'.format(meta_samples_df.loc[item,:].corr(grouped_series, method='spearman')) + '\t' + '{0:.3}'.format(meta_samples_df.loc[item,meta_samples_df.columns[1]:].corr(grouped_series[:-1], method='spearman')))
        text = 'spearman (rank) correlation' + ':\n'
        corr_series = pd.Series(corr_list, index=meta_samples_df.index)
        return text, corr_series
    